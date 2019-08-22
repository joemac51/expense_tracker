import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook

class Program:
    '''Main program class'''
    
    def __init__(self):
        self.current_month = datetime.now().strftime('%B') # the current month
        self.principal = 0 # starting money
        self.income = 0 # income for the current month
        self.bills = 0 # bills for the current month
        self.file_altered = False
        
    def calc_gross_profit(self):
        '''Sets the principal for the current month'''
        
        self.principal = int(input(f'Please enter your principal for {self.current_month}:\n'))
        self.income = int(input(f'Please enter your income for {self.current_month}:\n'))
        self.bills = int(input(f'Please enter your bills for {self.current_month}:\n'))
        
        net_income = self.income - self.bills
        
        return self.principal+net_income
        
    def mem_file(self):
        '''Alters a txt file that has already been initialised with the text 'August\n1' in order to
        remember what month the program left off on'''
        
        file_path = 'D:\\Documents\\Python\\expense_tracker\\month_mem.txt'
        
        a_file = open(file_path)
        a_str = a_file.read()
        a_file.close()
        a_str = a_str.split()
        
        if a_str[0] != self.current_month: # If its a different month to the last time the program was accessed
            a_str[0] = self.current_month
            a_str[1] = int(a_str[1]) + 1 # add one to the counter
            a_str[1] = str(a_str[1])
            
            a_file = open(file_path, 'w')
            a_file.write(a_str[0]+'\n'+a_str[1])
            
            self.file_altered = True
            
    def _get_mem_file(self):
        '''Gets the current values from 'month_mem.txt' '''
        
        file_path = 'D:\\Documents\\Python\\expense_tracker\\month_mem.txt'
        
        a_file = open(file_path)
        a_str = a_file.read()
        a_file.close()
        
        return a_str.split()
        
    def set_xl_file(self,profit):
        '''Writes the current gross profit and month to the excel file using openpyxl'''
        
        wb = Workbook()
        ws = wb.active
        
        mem_arr = self._get_mem_file()
        
        ws.cell(int(mem_arr[1]),1,mem_arr[0]) # adds current month to the workbook
        ws.cell(int(mem_arr[1]),2,profit) # adds current month's profit to the workbook
        
        wb.save('month_profit_data.xlsx')
    
    def plot_profits(self):
        '''Plots gross profits vs. months'''
        
#       plt.xticks([i for i in range(12)],months)
#       plt.plot([i for i in range(12)],monthly_total)
#       plt.show()
        pass
    
if __name__ == '__main__':
    
    main = Program()
    
    main.mem_file()
    
    if main.file_altered:
        gross_profit = main.calc_gross_profit()
        main.set_xl_file(gross_profit)
    else:
        print('\nMonth already logged!')